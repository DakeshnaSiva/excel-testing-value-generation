import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from flask import Flask, request, jsonify

app = Flask(__name__)

def contains_special_characters(string):
    special_characters = set("!@#$%^&*()_+[]{}|;':\",.<>?")
    return any(char in special_characters for char in string)

def process_data_entry(entry):
    name = entry.get("name", "")
    url = entry.get("url", "")
    
    if contains_special_characters(name):
        return {"name": name, "url": url, "error": 400, "error message": "not valid"}
    else:
        return {"name": name, "url": url, "error": 200, "error message": "valid"}

@app.route('/process_data', methods=['POST'])
def process_data():
    try:
        json_data = request.get_json()
        result_data = [process_data_entry(entry) for entry in json_data]
        
        output_file = "kh.xlsx"
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        
        # Merge cells for the title
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "CHANGE POND TESTING BUGS"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid") 
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)  # Adjust the column count
        
        df = pd.DataFrame(result_data, columns=["name", "url", "error", "error message"])
        
        # DataFrame columns starting from A3
        for col_idx, column in enumerate(df.columns, start=1):
            col_cell = sheet.cell(row=3, column=col_idx)
            col_cell.value = column
            col_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            col_cell.font = Font(bold=True) 
        
        # DataFrame values starting from A4
        for row_idx, row in enumerate(df.values, start=4):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
        
        error_format = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  
        success_format = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for row_index, row_data in enumerate(result_data, start=4):
            error_code = row_data["error"]
            cell = sheet.cell(row=row_index, column=3)  
            
            if error_code == 200:
                cell.fill = success_format  
            elif error_code == 400:
                cell.fill = error_format   
    
        workbook.save(output_file)
        
        return jsonify({"message": "Data processed and saved"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
