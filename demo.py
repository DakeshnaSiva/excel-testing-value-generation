import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from flask import Flask, request, jsonify
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import re

app = Flask(__name__)

def contains_special_characters(string):
    special_characters = set("!@#$%^&*()_+[]{}|;':\",.<>?")
    return any(char in special_characters for char in string)

def is_valid_password(password):
    if len(password) >= 8:
        return True
    return False

def is_valid_name(name):
    if re.search(r'[A-Z]', name) and re.search(r'\d+', name) and contains_special_characters(name):
        return True
    return False

def process_data_entry(entry, test_case_number):
    name = entry.get("name", "")
    password = entry.get("password", "")
    
    is_name_valid = is_valid_name(name)
    is_password_valid = is_valid_password(password)
    
    if is_name_valid and is_password_valid:
        response_message = "message: message was successfully posted"
        error_code = "successful status code of 200"
        test_case_number = f"TestCase{test_case_number}_Valid_Name_Password"
    else:
        response_message = "message: message was unsuccessful and not posted"
        error_code = "error: unsuccessful status code of 422"
        test_case_number = f"TestCase{test_case_number}_IN_Valid_Name_Password"

    request_text = f"{{attribute_name}}:{name} and {{attribute_code}}:{password}"
    
    return {
        "Test Case Number": test_case_number,
        "Request": request_text,
        "Response": response_message,
        "Response Message": error_code
    }

@app.route('/process_data', methods=['POST', 'GET'])
def process_data():
    try:
        json_data = request.get_json()
        result_data = []

        for idx, entry in enumerate(json_data, start=1):
            result_data.append(process_data_entry(entry, idx))

        output_file = "ressu.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"

        # Merge cells for the title
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "CHANGE POND TESTING BUGS"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)  # Adjust the column count

        df = pd.DataFrame(result_data, columns=["Test Case Number", "Request", "Response", "Response Message"])

        for col_idx, column in enumerate(df.columns, start=1):
            col_cell = sheet.cell(row=2, column=col_idx)
            col_cell.value = column
            col_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            col_cell.font = Font(bold=True)

        for row_idx, row_data in enumerate(result_data, start=3):
            for col_idx, value in enumerate(row_data.values(), start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value

        error_format = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        success_format = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for row_index, row_data in enumerate(result_data, start=3):
            error_code = row_data["Response Message"]
            cell = sheet.cell(row=row_index, column=3)  # Adjust the column index for the "Response" column

            if error_code == "successful status code of 200":
                cell.fill = success_format
            else:
                cell.fill = error_format

        workbook.save(output_file)

        return jsonify({"message": result_data}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 422

if __name__ == '__main__':
    app.run(debug=True)
