import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from flask import Flask, request, jsonify
import re
import json

app = Flask(__name__)

def contains_special_characters(string):
    special_characters = set("!@#$%^&*()_+[]{}|;':\",.<>?")
    return any(char in special_characters for char in string)

def is_valid_password(password):
    if isinstance(password, str):
        return len(password) >= 8
    return False

def is_valid_name(name):
    if isinstance(name, str):
        return bool(re.search(r'[A-Z]', name) and contains_special_characters(name))
    return False

def process_data_entry(entry, test_case_number, is_extracting=False):
    def flatten_dict(d, parent_key='', sep='_'):
        items = {}
        for k, v in d.items():
            new_key = parent_key + sep + k if parent_key else k
            if isinstance(v, dict):
                items.update(flatten_dict(v, new_key, sep=sep))
            elif isinstance(v, list):
                for i, item in enumerate(v):
                    if isinstance(item, (dict, list)):
                        # Recursively flatten nested dictionaries or lists within the list
                        items.update(flatten_dict({f'{new_key}_{i}': item}, sep=sep))
                    else:
                        items[f'{new_key}_{i}'] = item
            else:
                items[new_key] = v
        return items

    result = {}
    
    if isinstance(entry, (dict, list)):
        flattened_data = flatten_dict({'data': entry})
        request_data = flattened_data.get('data', {})
        name = request_data.get('name', '')

        if is_extracting:
            result["request"] = json.dumps(request_data, indent=4)
        else:
            if is_valid_name(name):
                result["Test Case Number"] = f"TestCase{test_case_number}_Valid_Entry"
                result["Request"] = json.dumps(request_data, indent=4)
                result["Response"] = "message: message was successfully posted"
                result["Response Message"] = "successful status code of 200"
            else:
                result["Test Case Number"] = f"TestCase{test_case_number}_Invalid_Entry"
                result["Request"] = json.dumps(entry, indent=4)
                result["Response"] = "message: message was unsuccessful and not posted"
                result["Response Message"] = "error: Invalid entry"

    return result

@app.route('/process_data', methods=['POST', 'GET'])
def process_data():
    try:
        json_data = request.get_json()
        
        if json_data is None:
            return jsonify({"error": "Invalid JSON data"}), 400

        result_data = []
        all_valid = True  # Flag to check if all entries are valid

        if isinstance(json_data, dict) and 'agencyCustomers' in json_data:
            agency_customers = json_data.get('agencyCustomers', [])
            if not isinstance(agency_customers, list):
                return jsonify({"error": "'agencyCustomers' must be a list"}), 400
        elif isinstance(json_data, list):
            agency_customers = json_data
        else:
            return jsonify({"error": "JSON data must be a list or a dictionary containing 'agencyCustomers'"}), 400

        for idx, entry in enumerate(agency_customers, start=1):
            result = process_data_entry(entry, idx, is_extracting=False)
            result_data.append(result)

            if "error" in result.get("Response Message", ""):
                all_valid = False

        output_file = "lox.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"

        # Merge cells for the title
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "CHANGE POND TESTING BUGS"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)  # Adjust the column count

        df = pd.DataFrame(result_data, columns=["Test Case Number", "Request", "Response", "Response Message"])

        for col_idx, column in enumerate(df.columns, start=1):
            col_cell = sheet.cell(row=2, column=col_idx)
            col_cell.value = column
            col_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
            col_cell.font = Font(bold=True)

        for row_idx, row_data in enumerate(result_data, start=3):
            for col_idx, value in enumerate(row_data.values(), start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value

        error_format = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        success_format = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for row_index, row_data in enumerate(result_data, start=3):
            error_code = row_data.get("Response Message", "")
            cell = sheet.cell(row=row_index, column=3)  # Adjust the column index for the "Response" column

            if "error" in error_code:
                cell.fill = error_format
            else:
                cell.fill = success_format

        workbook.save(output_file)

        return jsonify({"message": result_data})

    except Exception as e:
        return jsonify({"error": str(e)}), 422

if __name__ == '__main__':
    app.run(debug=True)
